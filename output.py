import openpyxl
import xlwings as xw

def excel(name,f_number, diff_percent,split,len_bs,res,sheet):
    wb = openpyxl.load_workbook("Input/"+name+'.xlsx', read_only=False)
    ws=wb[sheet]
    row_count = ws.max_row+1
    for i in range(split):
        r = row_count + i
        ws.cell(r, 1).value = str(f_number)
        ws.cell(r, 2).value = diff_percent
        ws.cell(r, 3).value = i
        for j in range(len_bs):
            try:
                ws.cell(r, 4 + j*3).value = res[j][i][1][1]
                ws.cell(r, 5 + j*3).value = res[j][i][0][0]
                ws.cell(r, 6 + j * 3).value = str(res[j][i][2])
            except:
                ws.cell(r, 4 + j * 3).value = 0
                ws.cell(r, 5 + j * 3).value = 0
                ws.cell(r, 6 + j * 3).value = "ERROR"
    r = ws.max_row + 1
    for line,item in zip(range(3),['MIN','MAX','AVERAGE']):
        l=r+line
        ws['D'+ str(l)] = '=' +item+'(D'+str(r-5)+':D'+str(r-1)+')'
        ws['E' + str(l)] = '=' +item+'(E'+str(r-5)+':E' + str(r - 1) + ')'
        ws['G' + str(l)] = '=' +item+'(G'+str(r-5)+':G' + str(r - 1) + ')'
        ws['H' + str(l)] = '=' +item+'(H'+str(r-5)+':H' + str(r - 1) + ')'
        ws['J' + str(l)] = '=' + item + '(J' + str(r - 5) + ':J' + str(r - 1) + ')'
        ws['K' + str(l)] = '=' +item+'(K'+str(r-5)+':K' + str(r - 1) + ')'
        #if 4 smells
        ws['M' + str(l)] = '=' + item + '(M' + str(r - 5) + ':M' + str(r - 1) + ')'
        ws['N' + str(l)] = '=' + item + '(N' + str(r - 5) + ':N' + str(r - 1) + ')'
    wb.save("Input/"+name+".xlsx")

def color(name,sheet_range):
    excel_app = xw.App(visible=True)
    wb = excel_app.books.open("Input/" + name + '.xlsx')
    sheet_i=0
    for item in wb.sheets:
        if sheet_i >= sheet_range:
            break
        sheet_i+=1
        min=[1,1,1,1,1,1,1,1]
        max=[0,0,0,0,0,0,0,0]
        row_min = [0, 0, 0, 0, 0, 0,0,0]
        row_max = [0, 0, 0, 0, 0, 0,0,0]
        col = ['D', 'E', 'G', 'H', 'J', 'K','M','N']
        '''min = [1, 1, 1, 1, 1, 1]
        max = [0, 0, 0, 0, 0, 0]
        row_min = [0, 0, 0, 0, 0, 0]
        row_max = [0, 0, 0, 0, 0, 0]
        col = ['D', 'E', 'G', 'H', 'J', 'K', 'M', 'N']'''
        #LastRow = xw.Range('A1').last_cell.row_max
        for i in range(9,100,8):
            for c in range(len(col)):
                print(item[col[c]+str(i)].value)
                if item[col[c]+str(i)].value==None:
                    break
                elif item[col[c]+str(i)].value>max[c]:
                    row_max[c]=i
                    max[c]=item[col[c]+str(i)].value
                elif item[col[c]+str(i)].value<min[c]:
                    row_min[c]=i
                    min[c]=item[col[c]+str(i)].value
        for j in range(len(col)):
            try:
                item[col[j] + str(row_min[j])].color = (255, 0, 0)
                item[col[j]+str(row_max[j])].color = (0,255,0)
            except:
                continue
    wb.save("Input/"+name+'.xlsx')
    wb.close()
    excel_app.quit()


