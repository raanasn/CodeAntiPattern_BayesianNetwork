from openpyxl import *

def excel(name,f_number, diff_percent,split,len_bs,res):
    wb = load_workbook("Input/"+name+'.xlsx', read_only=False)
    ws=wb["Sheet1"]
    row_count = ws.max_row
    for i in range(split):
        r = row_count + i
        ws.cell(r, 1).value = str(f_number)
        ws.cell(r, 2).value = diff_percent
        ws.cell(r, 3).value = i
        for j in range(len_bs):
            #print(res[j][i])
            ws.cell(r, 4 + j*2).value = res[j][i][1][1]
            ws.cell(r, 5 + j*2).value = res[j][i][0][0]
    r = ws.max_row + 1
    for line,item in zip(range(3),['MIN','MAX','AVERAGE']):
        l=r+line
        ws['D'+ str(l)] = '=' +item+'(D1:D'+str(r-1)+')'
        ws['E' + str(l)] = '=' +item+'(E1:E' + str(r - 1) + ')'
        ws['F' + str(l)] = '=' +item+'(F1:F' + str(r - 1) + ')'
        ws['G' + str(l)] = '=' +item+'(G1:G' + str(r - 1) + ')'
        ws['H' + str(l)] = '=' +item+'(H1:H' + str(r - 1) + ')'
        ws['I' + str(l)] = '=' +item+'(I1:I' + str(r - 1) + ')'
    wb.save("Input/"+name+".xlsx")